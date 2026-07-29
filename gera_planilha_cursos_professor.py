from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PRIMARY, SURFACE2, BORDER = "32327F", "EDF2F7", "E2E8F0"
F = "Arial"
th_font = Font(name=F, size=9, bold=True, color="FFFFFF")
th_fill = PatternFill("solid", fgColor=PRIMARY)
tot_font = Font(name=F, size=10, bold=True, color="2B2B32")
tot_fill = PatternFill("solid", fgColor=SURFACE2)
cel_font = Font(name=F, size=10, color="565758")
sub_font = Font(name=F, size=9, color="808285")
mut_font = Font(name=F, size=9, color="808285")
h1_font = Font(name=F, size=14, bold=True, color=PRIMARY)
bd = Border(bottom=Side(style="thin", color=BORDER))

BRL, PCT, INT = 'R$ #,##0.00', '0.00%', '#,##0'

wb = Workbook()

# ------------------------------------------------------------- Midia por turma
ws = wb.active
ws.title = "Mídia por turma"
ws["A1"] = "Cursos com turmas do Professor | Mídia paga por turma e plataforma"
ws["A1"].font = h1_font
ws["A2"] = "Período: 01/01/2026 a 30/06/2026"
ws["A2"].font = mut_font

cab = ["Curso", "Turma", "Plataforma", "Campanhas", "Janela", "Impressões",
       "Cliques", "CTR", "CPC", "Investimento", "Leads plataforma", "Leads RD", "CPL"]
for i, c in enumerate(cab, 1):
    cel = ws.cell(row=4, column=i, value=c)
    cel.font = th_font
    cel.fill = th_fill
    cel.alignment = Alignment(horizontal="right" if i > 5 else "left", vertical="center", wrap_text=True)

# curso, [(turma, plataforma, campanhas, janela, impressoes, cliques, custo, leads_plat)], leads_rd
dados = [
    ("Green Belt Lean Seis Sigma", [
        ("T101", "Meta", 2, "05/01 a 23/01", 142391, 1633, 2424.02, 226),
        ("T102", "Meta", 2, "27/01 a 13/03", 277844, 2800, 3704.34, 442),
        ("T103", "Meta", 1, "17/04 a 26/05", 186700, 2639, 4162.72, 387),
        ("T103", "LinkedIn", 1, "07/05 a 08/05", 1713, 12, 277.60, 0),
        ("T99 continuada", "LinkedIn", 1, "28/04 a 30/04", 2303, 16, 195.11, 2),
        ("T104", "Meta", 1, "10/06 a 30/06", 45168, 715, 1247.44, 138),
        ("T104", "LinkedIn", 1, "18/06 a 30/06", 13415, 132, 420.23, 32),
    ], 1587),
    ("Upgrade para Black Belt em Lean Seis Sigma", [
        ("CCUPON T42", "Meta", 2, "12/01 a 31/03", 167304, 2140, 2714.75, 213),
        ("CCUPON T42", "LinkedIn", 1, "24/02 a 26/03", 63387, 502, 2002.78, 100),
    ], 392),
    ("Yellow Belt Lean Seis Sigma", [
        ("T26", "Meta", 2, "05/01 a 30/01", 50542, 821, 962.45, 64),
        ("T27", "Meta", 1, "04/03 a 29/03", 55432, 748, 1268.52, 81),
        ("T28", "Meta", 3, "18/05 a 10/06", 42416, 999, 1363.35, 114),
    ], 332),
    ("Experimentos na Prática para Melhoria de Produtos e Processos com JASP", [
        ("T1", "Meta", 1, "24/03 a 07/05", 209914, 3207, 1926.33, 350),
        ("T1", "LinkedIn", 1, "24/03 a 25/03", 875, 3, 70.59, 0),
    ], 333),
    ("Lean: Ferramentas para Excelência Operacional", [
        ("T25", "Meta", 2, "11/05 a 19/06", 114932, 1763, 1078.64, 157),
    ], 206),
]

r = 5
linhas_total = []
for curso, turmas, leads_rd in dados:
    ini = r
    for turma, plat, camp, janela, impr, cliq, custo, lp in turmas:
        ws.cell(row=r, column=1, value=curso).font = sub_font
        ws.cell(row=r, column=2, value=turma).font = cel_font
        ws.cell(row=r, column=3, value=plat).font = cel_font
        ws.cell(row=r, column=4, value=camp).font = cel_font
        ws.cell(row=r, column=5, value=janela).font = cel_font
        ws.cell(row=r, column=6, value=impr).font = cel_font
        ws.cell(row=r, column=7, value=cliq).font = cel_font
        ws.cell(row=r, column=8, value=f"=IFERROR(G{r}/F{r},0)").font = cel_font
        ws.cell(row=r, column=9, value=f"=IFERROR(J{r}/G{r},0)").font = cel_font
        ws.cell(row=r, column=10, value=custo).font = cel_font
        ws.cell(row=r, column=11, value=lp).font = cel_font
        r += 1
    fim = r - 1
    ws.cell(row=r, column=1, value=curso).font = tot_font
    ws.cell(row=r, column=2, value="Total do curso").font = tot_font
    for col in (4, 6, 7, 10, 11):
        L = get_column_letter(col)
        ws.cell(row=r, column=col, value=f"=SUM({L}{ini}:{L}{fim})").font = tot_font
    ws.cell(row=r, column=8, value=f"=IFERROR(G{r}/F{r},0)").font = tot_font
    ws.cell(row=r, column=9, value=f"=IFERROR(J{r}/G{r},0)").font = tot_font
    ws.cell(row=r, column=12, value=leads_rd).font = tot_font
    ws.cell(row=r, column=13, value=f"=IFERROR(J{r}/L{r},0)").font = tot_font
    for col in range(1, 14):
        ws.cell(row=r, column=col).fill = tot_fill
    linhas_total.append(r)
    r += 1

rg = r
ws.cell(row=rg, column=1, value="Total geral")
ws.cell(row=rg, column=2, value="")
for col in (4, 6, 7, 10, 11, 12):
    L = get_column_letter(col)
    ws.cell(row=rg, column=col, value="=SUM(" + ",".join(f"{L}{x}" for x in linhas_total) + ")")
ws.cell(row=rg, column=8, value=f"=IFERROR(G{rg}/F{rg},0)")
ws.cell(row=rg, column=9, value=f"=IFERROR(J{rg}/G{rg},0)")
ws.cell(row=rg, column=13, value=f"=IFERROR(J{rg}/L{rg},0)")
for col in range(1, 14):
    c = ws.cell(row=rg, column=col)
    c.font = Font(name=F, size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=PRIMARY)

for row in range(5, rg + 1):
    for col, fmt in ((4, INT), (6, INT), (7, INT), (8, PCT), (9, BRL), (10, BRL), (11, INT), (12, INT), (13, BRL)):
        ws.cell(row=row, column=col).number_format = fmt

notas = [
    "A T104 do Green Belt seguia no ar em 30/06, então os números dela param na data de corte do relatório.",
    "Leads plataforma é o que Meta e LinkedIn reportam. Leads RD é pessoa única por curso pela régua anti-refire de 90 dias, e por isso só existe no total do curso.",
    "Fontes: Meta Ads e LinkedIn Ads via Reportei; RD Station para leads. CTR, CPC e CPL calculados na planilha.",
]
for i, t in enumerate(notas):
    ws.cell(row=rg + 2 + i, column=1, value=t).font = mut_font

for i, w in enumerate([46, 15, 11, 11, 15, 13, 10, 9, 10, 14, 15, 10, 10], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ------------------------------------------------------- Posicionamento SEO
ws2 = wb.create_sheet("Posicionamento orgânico")
ws2["A1"] = "Posicionamento orgânico nos temas da linha"
ws2["A1"].font = h1_font
ws2["A2"] = "Consulta de julho de 2026"
ws2["A2"].font = mut_font
for i, c in enumerate(["Tema", "Volume mensal", "Dificuldade", "Vanzolini", "Melhor concorrente posicionado", "Diagnóstico"], 1):
    cel = ws2.cell(row=4, column=i, value=c)
    cel.font = th_font
    cel.fill = th_fill
    cel.alignment = Alignment(horizontal="right" if i in (2, 3) else "left", vertical="center", wrap_text=True)
seo = [
    ("Lean Six Sigma", 5400, 23, "2ª", "FM2S (1ª)", "Paridade competitiva"),
    ("Black Belt", 5400, 17, "8ª", "FM2S (1ª)", "Desvantagem competitiva"),
    ("Green Belt", 4400, 16, "3ª", "FM2S (9ª)", "Liderança consolidada"),
    ("DMAIC", 3600, 19, "Não posicionada", "FM2S (2ª)", "Ausência de cobertura"),
    ("JASP", 1900, 40, "Não posicionada", "Sem concorrente setorial", "Intenção incompatível"),
    ("Melhoria contínua", 1300, 24, "4ª", "Sem concorrente no top 10", "Cobertura indireta"),
    ("Excelência operacional", 720, 17, "Não posicionada", "FM2S (6ª)", "Ausência de cobertura"),
    ("IA aplicada à melhoria contínua", "Sem dados", "Não aplicável", "Não aplicável", "Não aplicável", "Demanda inexistente"),
]
r = 5
for linha in seo:
    for i, v in enumerate(linha, 1):
        c = ws2.cell(row=r, column=i, value=v)
        c.font = cel_font
        c.border = bd
        if i in (2, 3) and isinstance(v, int):
            c.number_format = INT
    r += 1
ws2.cell(row=r + 1, column=1, value="Volume mensal e dificuldade de palavra-chave para o Brasil. Posição orgânica no Google.").font = mut_font
for i, w in enumerate([32, 14, 12, 18, 28, 24], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A5"

# ---------------------------------------------------------- Guia Lean 6 Sigma
ws3 = wb.create_sheet("Guia Lean Seis Sigma")
ws3["A1"] = "Guia Lean Seis Sigma"
ws3["A1"].font = h1_font
ws3["A2"] = "vanzolini.org.br/blog/lean-seis-sigma-estruturacao-precisao-e-impacto/"
ws3["A2"].font = Font(name=F, size=10, color=PRIMARY, underline="single")
ws3["A2"].hyperlink = "https://vanzolini.org.br/blog/lean-seis-sigma-estruturacao-precisao-e-impacto/"
for i, c in enumerate(["Indicador", "Valor"], 1):
    cel = ws3.cell(row=4, column=i, value=c)
    cel.font = th_font
    cel.fill = th_fill
    cel.alignment = Alignment(horizontal="right" if i == 2 else "left", vertical="center")
guia = [
    ("Pessoas captadas pelo guia, de abril de 2025 a julho de 2026", 22, INT),
    ("Conversões brutas registradas do guia no mesmo período", 59, INT),
    ("Conversões do guia em 2026", 15, INT),
    ("Pessoas captadas pelo pop-up do guia, no ar em 13 e 14/08/2025", 13, INT),
    ("Total de pessoas captadas pelo guia e pelo pop-up", None, INT),
    ("Dessas pessoas, quantas converteram depois em algum curso", 7, INT),
    ("Cursos diferentes atingidos por essas pessoas", 7, INT),
    ("Cliques orgânicos da página do guia no semestre", 59, INT),
    ("Cliques orgânicos das 28 páginas da linha Lean Seis Sigma", 4914, INT),
    ("Participação da página do guia no tráfego da linha", None, '0.0%'),
]
r = 5
for nome, valor, fmt in guia:
    ws3.cell(row=r, column=1, value=nome).font = cel_font
    c = ws3.cell(row=r, column=2)
    if nome.startswith("Total de pessoas"):
        c.value = "=B5+B8"
    elif nome.startswith("Participação"):
        c.value = "=IFERROR(B12/B13,0)"
    else:
        c.value = valor
    c.font = cel_font
    c.number_format = fmt
    ws3.cell(row=r, column=1).border = bd
    c.border = bd
    r += 1
ws3.cell(row=r + 1, column=1, value="Fontes: RD Station para conversões e Search Console para cliques orgânicos do semestre.").font = mut_font
ws3.column_dimensions["A"].width = 62
ws3.column_dimensions["B"].width = 14


# --------------------------------------------------------- Turmas e matriculas
ws4 = wb.create_sheet("Turmas e matrículas")
ws4["A1"] = "Turmas de 2026 dos cursos do Professor | inscritos, matrículas e custo por matrícula"
ws4["A1"].font = h1_font
ws4["A2"] = "Mídia de 01/01 a 30/06/2026. Inscritos e matrículas conforme a planilha GESTÃO VANZOLINI 2025/2026, aba Comercial_Status."
ws4["A2"].font = mut_font

cab4 = ["Curso", "Turma", "Início", "Status", "Mídia no período", "Investimento",
        "Leads plataforma", "Inscritos", "Matrículas pagantes", "Custo por matrícula"]
for i, c in enumerate(cab4, 1):
    cel = ws4.cell(row=4, column=i, value=c)
    cel.font = th_font
    cel.fill = th_fill
    cel.alignment = Alignment(horizontal="right" if i > 5 else "left", vertical="center", wrap_text=True)

# curso, turma, inicio, status, teve_midia, investimento, leads_plat, inscritos, matriculas
tur = [
    ("Green Belt Lean Seis Sigma", "CCGBON T 101", "26/01/2026", "Realizado", "Sim", 2424.02, 226, 60, 42),
    ("Green Belt Lean Seis Sigma", "CCGBON T 102", "07/04/2026", "Realizado", "Sim", 3704.34, 442, 71, 41),
    ("Green Belt Lean Seis Sigma", "CCGBON T 103", "01/06/2026", "Realizado", "Sim", 4440.32, 387, 63, 44),
    ("Green Belt Lean Seis Sigma", "CCGBON T 104", "11/08/2026", "Realizado", "Sim", 1667.67, 170, 69, 44),
    ("Green Belt Lean Seis Sigma", "CCGBON T 99 (turma de 2025)", "30/10/2025", "Realizado", "Sim", 195.11, 2, None, None),
    ("Green Belt Lean Seis Sigma", "CCGBON T 105", "24/09/2026", "A cadastrar", "Não", None, None, 27, 3),
    ("Green Belt Lean Seis Sigma", "CCGBON T 106", "27/10/2026", "A definir", "Não", None, None, None, None),
    ("Upgrade para Black Belt em Lean Seis Sigma", "CCUPON T 42", "02/04/2026", "Realizado", "Sim", 4717.53, 313, 32, 21),
    ("Upgrade para Black Belt em Lean Seis Sigma", "CCUPON T 43", "21/09/2026", "Planejado", "Não", None, None, 12, 5),
    ("Yellow Belt Lean Seis Sigma", "A-YBON T 26", "02/02/2026", "Realizado", "Sim", 962.45, 64, 16, 15),
    ("Yellow Belt Lean Seis Sigma", "A-YBON T 27", "07/04/2026", "Realizado", "Sim", 1268.52, 81, 15, 14),
    ("Yellow Belt Lean Seis Sigma", "A-YBON T 28", "08/06/2026", "Realizado", "Sim", 1363.35, 114, 19, 16),
    ("Yellow Belt Lean Seis Sigma", "A-YBON T 29", "17/08/2026", "Planejado", "Não", None, None, 22, 15),
    ("Yellow Belt Lean Seis Sigma", "A-YBON T 30", "06/10/2026", "Planejado", "Não", None, None, 0, 0),
    ("Experimentos na Prática com JASP", "A-EPMPPON T 01", "18/05/2026", "Descontinuado", "Sim", 1996.92, 350, 0, 0),
    ("Lean: Ferramentas para Excelência Operacional", "A-LMFON T 24", "23/03/2026", "Realizado", "Não", None, None, 9, 7),
    ("Lean: Ferramentas para Excelência Operacional", "A-LMFON T 25", "01/07/2026", "Cancelada", "Sim", 1078.64, 157, 2, 0),
    ("Lean: Ferramentas para Excelência Operacional", "A-LMFON T 26", "21/09/2026", "Planejado", "Não", None, None, 0, 0),
]
r = 5
com_midia = []
for curso, turma, ini, st, mid, inv, lp, ins, mat in tur:
    ws4.cell(row=r, column=1, value=curso).font = cel_font
    ws4.cell(row=r, column=2, value=turma).font = cel_font
    ws4.cell(row=r, column=3, value=ini).font = cel_font
    ws4.cell(row=r, column=4, value=st).font = cel_font
    ws4.cell(row=r, column=5, value=mid).font = cel_font
    ws4.cell(row=r, column=6, value=inv).font = cel_font
    ws4.cell(row=r, column=7, value=lp).font = cel_font
    ws4.cell(row=r, column=8, value=ins).font = cel_font
    ws4.cell(row=r, column=9, value=mat).font = cel_font
    ws4.cell(row=r, column=10, value=f"=IF(AND(F{r}<>\"\",I{r}>0),F{r}/I{r},\"\")").font = cel_font
    for col in range(1, 11):
        ws4.cell(row=r, column=col).border = bd
    if mid == "Sim":
        com_midia.append(r)
    r += 1

ws4.cell(row=r, column=1, value="Total das turmas com mídia no período").font = Font(name=F, size=10, bold=True, color="FFFFFF")
for col in (6, 7, 8, 9):
    L = get_column_letter(col)
    ws4.cell(row=r, column=col, value="=SUM(" + ",".join(f"{L}{x}" for x in com_midia) + ")")
ws4.cell(row=r, column=10, value=f"=IFERROR(F{r}/I{r},0)")
for col in range(1, 11):
    c = ws4.cell(row=r, column=col)
    c.font = Font(name=F, size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=PRIMARY)

for row in range(5, r + 1):
    for col, fmt in ((6, BRL), (7, INT), (8, INT), (9, INT), (10, BRL)):
        ws4.cell(row=row, column=col).number_format = fmt

notas4 = [
    "Matrículas pagantes é a coluna Matriculas pgtes da aba Comercial_Status. Inscritos é a coluna INSCRITOS da mesma aba.",
    "A T104 do Green Belt seguiu com mídia depois de 30/06, então o investimento dela na tabela é parcial e o custo por matrícula está subestimado.",
    "A T99 do Green Belt é turma de 2025 que recebeu mídia de continuada em abril de 2026, por isso não tem inscritos nem matrículas de 2026.",
]
for i, t in enumerate(notas4):
    ws4.cell(row=r + 2 + i, column=1, value=t).font = mut_font

for i, w in enumerate([44, 26, 12, 14, 14, 14, 15, 11, 17, 16], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = "A5"


saida = "/private/tmp/claude-501/-Users-junior-Library-CloudStorage-GoogleDrive-jr-communitas-com-br-Meu-Drive-C-rebro-Communitas/ea2e053f-6aba-4236-9e68-a4be188c8040/scratchpad/vanzolini_midia_cursos_professor_janeiro_a_junho_2026.xlsx"
wb.save(saida)
print("salvo:", saida)
